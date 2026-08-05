"""The comparison workbook is the audit trail, so its layout is asserted literally."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from clearml_yolo.comparison.workbook import write_comparison_workbook

EXPECTED_HEADERS = [
    "Класс",
    "Порог прод",
    "Порог новой",
    "TP прод",
    "FP прод",
    "FN прод",
    "TP новая",
    "FP новая",
    "FN новая",
    "Precision прод",
    "Precision новая",
    "Δ Precision",
    "Δ CI низ (P)",
    "Δ CI верх (P)",
    "p (P)",
    "p BH (P)",
    "Вердикт (P)",
    "Recall прод",
    "Recall новая",
    "Δ Recall",
    "Δ CI низ (R)",
    "Δ CI верх (R)",
    "p (R)",
    "p BH (R)",
    "Вердикт (R)",
]


def _row(
    class_name: str,
    *,
    is_pooled: bool = False,
    precision_baseline: float = 0.5,
    precision_p_bh: float = 0.02,
    precision_verdict: str = "improved",
) -> dict[str, object]:
    return {
        "class_name": class_name,
        "is_pooled": is_pooled,
        "threshold_baseline": 0.35,
        "threshold_candidate": 0.4,
        "tp_baseline": 10,
        "fp_baseline": 3,
        "fn_baseline": 4,
        "tp_candidate": 12,
        "fp_candidate": 2,
        "fn_candidate": 2,
        "precision_baseline": precision_baseline,
        "precision_candidate": 0.857142857,
        "precision_delta": 0.0952380952,
        "precision_ci_lower": -0.0123456789,
        "precision_ci_upper": 0.2028398748,
        "precision_p_value": 0.0312345678,
        "precision_p_bh": precision_p_bh,
        "precision_verdict": precision_verdict,
        "recall_baseline": 0.714285714,
        "recall_candidate": 0.857142857,
        "recall_delta": 0.142857143,
        "recall_ci_lower": 0.0111111111,
        "recall_ci_upper": 0.2745555555,
        "recall_p_value": 0.0456789012,
        "recall_p_bh": 0.0912345678,
        "recall_verdict": "not_significant",
    }


@pytest.fixture
def rows() -> pd.DataFrame:
    return pd.DataFrame(
        [
            _row("total", is_pooled=True, precision_p_bh=float("nan")),
            _row("cat"),
            _row("dog", precision_baseline=float("nan"), precision_verdict="degraded"),
        ]
    )


@pytest.fixture
def excluded() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"class_name": "ferret", "reason": "нет GT-объектов в сплите"},
            {"class_name": "otter", "reason": "мало примеров в обучении"},
        ]
    )


@pytest.fixture
def methodology() -> dict[str, object]:
    return {
        "Тест для precision": "бутстрэп по изображениям",
        "Тест для recall": "тест Макнемара",
        "Размер семейства гипотез": 2,
        "q (FDR)": 0.1,
        "Зерно ГПСЧ": 12345,
        "Источник порогов": "прод-конфиг",
        "Источник весов": "clearml://task/abc123",
        "Изображений": 500,
        "Боксов GT": 1370,
    }


def _write(
    tmp_path: Path,
    rows: pd.DataFrame,
    excluded: pd.DataFrame,
    methodology: dict[str, object],
) -> dict[str, pd.DataFrame]:
    destination = tmp_path / "nested" / "comparison.xlsx"
    write_comparison_workbook(rows, excluded, methodology, destination)
    assert destination.is_file()
    return pd.read_excel(destination, sheet_name=None)


def test_sheet_names_and_headers(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    book = _write(tmp_path, rows, excluded, methodology)

    assert list(book) == ["Сравнение", "Исключённые классы", "Методика"]
    assert list(book["Сравнение"].columns) == EXPECTED_HEADERS
    assert list(book["Исключённые классы"].columns) == ["Класс", "Причина"]
    assert list(book["Методика"].columns) == ["Параметр", "Значение"]


def test_class_rows_are_rounded_and_flags_are_readable(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    sheet = _write(tmp_path, rows, excluded, methodology)["Сравнение"]
    cat = sheet.loc[sheet["Класс"] == "cat"].iloc[0]

    assert cat["Precision новая"] == pytest.approx(0.8571)
    assert cat["Δ CI низ (P)"] == pytest.approx(-0.0123)
    assert cat["p BH (P)"] == pytest.approx(0.02)
    assert cat["TP новая"] == 12
    assert cat["p (P)"] == pytest.approx(0.0312345678)


def test_pooled_row_is_last_and_outside_the_bh_family(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    sheet = _write(tmp_path, rows, excluded, methodology)["Сравнение"]

    assert list(sheet["Класс"]) == ["cat", "dog", "Итого"]
    pooled = sheet.iloc[-1]
    assert pooled["p BH (P)"] == "не применимо"
    assert pooled["p BH (R)"] == "не применимо"
    assert pooled["p (P)"] == pytest.approx(0.0312345678)
    assert pooled["Вердикт (P)"] == "улучшение"


def test_every_verdict_renders_in_russian(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    sheet = _write(tmp_path, rows, excluded, methodology)["Сравнение"]

    assert list(sheet["Вердикт (P)"]) == ["улучшение", "деградация", "улучшение"]
    assert list(sheet["Вердикт (R)"]) == ["не значимо"] * 3


def test_degraded_and_improved_verdicts_are_visually_distinct(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    destination = tmp_path / "coloured.xlsx"
    write_comparison_workbook(rows, excluded, methodology, destination)
    sheet = load_workbook(destination)["Сравнение"]

    verdict_column = EXPECTED_HEADERS.index("Вердикт (P)") + 1
    improved = sheet.cell(row=2, column=verdict_column)
    degraded = sheet.cell(row=3, column=verdict_column)
    not_significant = sheet.cell(row=2, column=EXPECTED_HEADERS.index("Вердикт (R)") + 1)

    assert improved.value == "улучшение"
    assert degraded.value == "деградация"
    assert improved.fill.fgColor.rgb != degraded.fill.fgColor.rgb
    assert improved.fill.fill_type == "solid"
    assert degraded.fill.fill_type == "solid"
    assert not_significant.fill.fill_type is None
    assert sheet.cell(row=len(rows) + 1, column=verdict_column).fill.fill_type == "solid"
    assert sheet.freeze_panes == "B2"
    assert sheet.cell(row=1, column=1).font.bold
    assert sheet.cell(row=len(rows) + 1, column=1).font.bold


def test_unknown_verdict_is_rejected(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    broken = rows.copy()
    broken.loc[0, "recall_verdict"] = "maybe"
    broken.loc[1, "precision_verdict"] = float("nan")

    with pytest.raises(ValueError, match="maybe"):
        write_comparison_workbook(broken, excluded, methodology, tmp_path / "broken.xlsx")


def test_missing_methodology_values_render_as_a_dash(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame
) -> None:
    methodology: dict[str, object] = {
        "Источник весов": None,
        "Зерно ГПСЧ": float("nan"),
        "Изображений": 500,
    }

    destination = tmp_path / "methodology.xlsx"
    write_comparison_workbook(rows, excluded, methodology, destination)
    dump = load_workbook(destination)["Методика"]

    assert [dump.cell(row=number, column=2).value for number in (2, 3, 4)] == ["—", "—", "500"]


def test_missing_values_never_render_as_nan_text(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    sheet = _write(tmp_path, rows, excluded, methodology)["Сравнение"]
    dog = sheet.loc[sheet["Класс"] == "dog"].iloc[0]

    assert dog["Precision прод"] == "—"
    assert "nan" not in [str(value).lower() for value in sheet.to_numpy().ravel()]


def test_excluded_and_methodology_round_trip(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    book = _write(tmp_path, rows, excluded, methodology)

    assert list(book["Исключённые классы"]["Класс"]) == ["ferret", "otter"]
    assert book["Исключённые классы"]["Причина"].iloc[0] == "нет GT-объектов в сплите"

    dump = book["Методика"]
    assert list(dump["Параметр"]) == list(methodology)
    assert list(dump["Значение"].astype(str)) == [str(value) for value in methodology.values()]


@pytest.mark.parametrize(
    "empty", [pd.DataFrame(), pd.DataFrame(columns=["class_name", "reason"])], ids=["bare", "typed"]
)
def test_empty_excluded_frame_still_gets_a_headed_sheet(
    tmp_path: Path, rows: pd.DataFrame, methodology: dict[str, object], empty: pd.DataFrame
) -> None:
    sheet = _write(tmp_path, rows, empty, methodology)["Исключённые классы"]

    assert list(sheet.columns) == ["Класс", "Причина"]
    assert sheet.empty


def test_excluded_frame_with_wrong_columns_is_rejected(
    tmp_path: Path, rows: pd.DataFrame, methodology: dict[str, object]
) -> None:
    drifted = pd.DataFrame([{"name": "ferret", "reason": "нет GT-объектов в сплите"}])

    with pytest.raises(ValueError, match="class_name"):
        write_comparison_workbook(rows, drifted, methodology, tmp_path / "broken.xlsx")


def test_tiny_p_values_survive_display_rounding(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    strong = rows.copy()
    strong.loc[1, "precision_p_value"] = 1e-9
    strong.loc[1, "precision_p_bh"] = 3e-9

    sheet = _write(tmp_path, strong, excluded, methodology)["Сравнение"]
    cat = sheet.loc[sheet["Класс"] == "cat"].iloc[0]

    assert cat["p (P)"] == pytest.approx(1e-9)
    assert cat["p BH (P)"] == pytest.approx(3e-9)


def test_inputs_are_not_mutated(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    rows_before = rows.copy(deep=True)
    excluded_before = excluded.copy(deep=True)
    methodology_before = dict(methodology)

    _write(tmp_path, rows, excluded, methodology)

    pd.testing.assert_frame_equal(rows, rows_before)
    pd.testing.assert_frame_equal(excluded, excluded_before)
    assert methodology == methodology_before


def test_missing_canonical_column_is_rejected(
    tmp_path: Path, rows: pd.DataFrame, excluded: pd.DataFrame, methodology: dict[str, object]
) -> None:
    with pytest.raises(ValueError, match="recall_p_bh"):
        write_comparison_workbook(
            rows.drop(columns=["recall_p_bh"]), excluded, methodology, tmp_path / "broken.xlsx"
        )
