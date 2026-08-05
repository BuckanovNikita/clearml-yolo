"""Render the baseline-versus-candidate comparison into a minimal Excel workbook.

Sheet titles, column headers and cell markers are Russian because the workbook is read by
stakeholders; they are data, not documentation.
"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import pandas as pd
from loguru import logger
from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

COMPARISON_SHEET = "Сравнение"
EXCLUDED_SHEET = "Исключённые классы"
METHODOLOGY_SHEET = "Методика"

POOLED_LABEL = "Итого"
NOT_APPLICABLE = "не применимо"
MISSING_VALUE = "—"

IMPROVED = "улучшение"
DEGRADED = "деградация"
NOT_SIGNIFICANT = "не значимо"

VERDICT_LABELS = {
    "improved": IMPROVED,
    "degraded": DEGRADED,
    "not_significant": NOT_SIGNIFICANT,
}

DISPLAY_DECIMALS = 4
_MIN_COLUMN_WIDTH = 10
_MAX_COLUMN_WIDTH = 40

POOLED_FLAG = "is_pooled"

COMPARISON_COLUMNS: tuple[tuple[str, str], ...] = (
    ("class_name", "Класс"),
    ("threshold_baseline", "Порог прод"),
    ("threshold_candidate", "Порог новой"),
    ("tp_baseline", "TP прод"),
    ("fp_baseline", "FP прод"),
    ("fn_baseline", "FN прод"),
    ("tp_candidate", "TP новая"),
    ("fp_candidate", "FP новая"),
    ("fn_candidate", "FN новая"),
    ("precision_baseline", "Precision прод"),
    ("precision_candidate", "Precision новая"),
    ("precision_delta", "Δ Precision"),
    ("precision_ci_lower", "Δ CI низ (P)"),
    ("precision_ci_upper", "Δ CI верх (P)"),
    ("precision_p_value", "p (P)"),
    ("precision_p_bh", "p BH (P)"),
    ("precision_verdict", "Вердикт (P)"),
    ("recall_baseline", "Recall прод"),
    ("recall_candidate", "Recall новая"),
    ("recall_delta", "Δ Recall"),
    ("recall_ci_lower", "Δ CI низ (R)"),
    ("recall_ci_upper", "Δ CI верх (R)"),
    ("recall_p_value", "p (R)"),
    ("recall_p_bh", "p BH (R)"),
    ("recall_verdict", "Вердикт (R)"),
)

EXCLUDED_COLUMNS: tuple[tuple[str, str], ...] = (
    ("class_name", "Класс"),
    ("reason", "Причина"),
)

_VERDICT_SOURCES = ("precision_verdict", "recall_verdict")
_VERDICT_HEADERS = ("Вердикт (P)", "Вердикт (R)")
_BH_HEADERS = ("p BH (P)", "p BH (R)")
# p-values span orders of magnitude, so display rounding would turn a strong result into 0.
_FULL_PRECISION_HEADERS = ("p (P)", "p (R)", *_BH_HEADERS)

_BOLD = Font(bold=True)
_VERDICT_FILLS = {
    IMPROVED: PatternFill(fill_type="solid", fgColor="FFC6EFCE"),
    DEGRADED: PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
}


def _comparison_sheet(rows: pd.DataFrame) -> pd.DataFrame:
    required = [POOLED_FLAG, *(name for name, _ in COMPARISON_COLUMNS)]
    missing = [name for name in required if name not in rows.columns]
    if missing:
        raise ValueError(f"Comparison rows are missing required columns: {missing}")

    verdicts = set(rows[_VERDICT_SOURCES[0]]) | set(rows[_VERDICT_SOURCES[1]])
    unknown = sorted(str(verdict) for verdict in verdicts - set(VERDICT_LABELS))
    if unknown:
        raise ValueError(f"Comparison rows carry unknown verdicts: {unknown}")

    ordered = rows.sort_values(POOLED_FLAG, kind="stable")
    pooled = ordered[POOLED_FLAG].astype(bool).to_numpy()
    frame = ordered.loc[:, [name for name, _ in COMPARISON_COLUMNS]].rename(
        columns=dict(COMPARISON_COLUMNS)
    )
    frame = frame.reset_index(drop=True)
    rounded = [header for header in frame.columns if header not in _FULL_PRECISION_HEADERS]
    frame[rounded] = frame[rounded].round(DISPLAY_DECIMALS)

    frame.loc[pooled, "Класс"] = POOLED_LABEL
    for header in _VERDICT_HEADERS:
        frame[header] = frame[header].map(VERDICT_LABELS)
    for header in _BH_HEADERS:
        frame[header] = frame[header].astype(object)
        frame.loc[pooled, header] = NOT_APPLICABLE
    return frame


def _style_sheet(worksheet: Any, frame: pd.DataFrame, bold_last_row: bool) -> None:
    worksheet.freeze_panes = "B2"
    for cell in worksheet[1]:
        cell.font = _BOLD

    for position, header in enumerate(frame.columns, start=1):
        widest_value = max((len(str(value)) for value in frame[header]), default=0)
        width = max(len(str(header)), widest_value) + 2
        letter = get_column_letter(position)
        worksheet.column_dimensions[letter].width = min(
            max(width, _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH
        )

    if bold_last_row and len(frame):
        for cell in worksheet[len(frame) + 1]:
            cell.font = _BOLD


def _rendered(value: object) -> str:
    """Stringify a methodology value without ever spelling a missing one as ``nan``."""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return MISSING_VALUE
    return str(value)


def _paint_verdicts(worksheet: Any, frame: pd.DataFrame) -> None:
    """Fill improved cells green and degraded cells red so regressions cannot be missed."""
    headers = list(frame.columns)
    for header in _VERDICT_HEADERS:
        column = headers.index(header) + 1
        for row_number, verdict in enumerate(frame[header], start=2):
            fill = _VERDICT_FILLS.get(str(verdict))
            if fill is not None:
                worksheet.cell(row=row_number, column=column).fill = fill


def write_comparison_workbook(
    rows: pd.DataFrame,
    excluded: pd.DataFrame,
    methodology: dict[str, object],
    path: Path,
) -> None:
    """Write the per-class comparison, the excluded classes and the methodology dump.

    ``rows`` carries the canonical English column names of :data:`COMPARISON_COLUMNS` plus
    ``is_pooled``; the pooled row is the micro-averaged hypothesis and stays outside the
    Benjamini-Hochberg family, so its corrected p-values are rendered as text. Neither input
    frame is modified.
    """
    comparison = _comparison_sheet(rows)
    unnamed = [name for name, _ in EXCLUDED_COLUMNS if name not in excluded.columns]
    if len(excluded) and unnamed:
        raise ValueError(f"Excluded classes are missing required columns: {unnamed}")
    excluded_sheet = excluded.reindex(columns=[name for name, _ in EXCLUDED_COLUMNS]).rename(
        columns=dict(EXCLUDED_COLUMNS)
    )
    methodology_sheet = pd.DataFrame(
        {
            "Параметр": [str(key) for key in methodology],
            "Значение": [_rendered(value) for value in methodology.values()],
        }
    )

    written = {
        COMPARISON_SHEET: comparison,
        EXCLUDED_SHEET: excluded_sheet,
        METHODOLOGY_SHEET: methodology_sheet,
    }
    pooled_row_present = bool(rows[POOLED_FLAG].astype(bool).any())

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for title, frame in written.items():
            frame.to_excel(writer, sheet_name=title, index=False, na_rep=MISSING_VALUE)
        for title, frame in written.items():
            bold_last_row = pooled_row_present and title == COMPARISON_SHEET
            _style_sheet(writer.sheets[title], frame, bold_last_row=bold_last_row)
        _paint_verdicts(writer.sheets[COMPARISON_SHEET], comparison)

    logger.info(
        "Comparison workbook: {} ({} classes, {} excluded)", path, len(comparison), len(excluded)
    )
